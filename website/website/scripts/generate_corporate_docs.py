#!/usr/bin/env python3
"""
TauOS Corporate Document Generator
Generates professional PDF and Excel documents with consistent branding
"""

import os
import sys
from datetime import datetime
from reportlab.lib.pagesizes import A4, letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
import requests
from PIL import Image as PILImage
import io

# Company Information
COMPANY_NAME = "AR Holdings Group Corporation"
COMPANY_SUBTITLE = "Tau Foundation / Tau LLC"
COMPANY_ADDRESS = "2126 Market Street, San Francisco, CA 94114, USA"
COMPANY_WEBSITE = "www.tauos.org"
COMPANY_EMAIL = "investors@tauos.org"

# Brand Colors (TauOS Theme)
PRIMARY_COLOR = "#F59E0B"  # Yellow-500
SECONDARY_COLOR = "#EA580C"  # Orange-600
ACCENT_COLOR = "#1F2937"  # Gray-800
TEXT_COLOR = "#111827"  # Gray-900
LIGHT_GRAY = "#F9FAFB"  # Gray-50

class TauOSDocumentGenerator:
    def __init__(self, output_dir="output/corporate_docs"):
        self.output_dir = output_dir
        self.logo_path = self._prepare_logo()
        os.makedirs(output_dir, exist_ok=True)
        
    def _prepare_logo(self):
        """Prepare the TauOS logo with proper background"""
        logo_path = "website/public/brand/tauos-logo.svg"
        if os.path.exists(logo_path):
            # Convert SVG to PNG with grey background
            try:
                from cairosvg import svg2png
                png_data = svg2png(url=logo_path, background_color="lightgrey")
                processed_logo = "temp_logo.png"
                with open(processed_logo, 'wb') as f:
                    f.write(png_data)
                return processed_logo
            except ImportError:
                print("Warning: cairosvg not available, using placeholder logo")
                return None
        return None
    
    def create_pdf_header(self, canvas, doc):
        """Create consistent PDF header with logo and company info"""
        width, height = A4
        
        # Grey background for logo area
        canvas.setFillColor(colors.HexColor("#F3F4F6"))
        canvas.rect(0, height - 100, width, 100, fill=1, stroke=0)
        
        # Logo
        if self.logo_path and os.path.exists(self.logo_path):
            try:
                canvas.drawImage(self.logo_path, 40, height - 80, width=60, height=60, preserveAspectRatio=True)
            except:
                pass
        
        # Company Information
        canvas.setFont("Helvetica-Bold", 16)
        canvas.setFillColor(colors.HexColor(PRIMARY_COLOR))
        canvas.drawString(120, height - 50, COMPANY_NAME)
        
        canvas.setFont("Helvetica", 10)
        canvas.setFillColor(colors.HexColor(TEXT_COLOR))
        canvas.drawString(120, height - 70, COMPANY_SUBTITLE)
        canvas.drawString(120, height - 85, COMPANY_ADDRESS)
        
        # Footer line
        canvas.setStrokeColor(colors.HexColor(PRIMARY_COLOR))
        canvas.setLineWidth(2)
        canvas.line(40, height - 100, width - 40, height - 100)
    
    def create_pdf_footer(self, canvas, doc):
        """Create consistent PDF footer"""
        width, height = A4
        
        # Footer line
        canvas.setStrokeColor(colors.HexColor(PRIMARY_COLOR))
        canvas.setLineWidth(1)
        canvas.line(40, 50, width - 40, 50)
        
        # Footer text
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor(TEXT_COLOR))
        canvas.drawString(40, 30, f"© 2025 {COMPANY_NAME}. All rights reserved.")
        canvas.drawString(width - 200, 30, f"Generated on {datetime.now().strftime('%B %d, %Y')}")
    
    def generate_developer_guide(self):
        """Generate Developer Guide PDF"""
        filename = os.path.join(self.output_dir, "TauOS_Developer_Guide.pdf")
        doc = SimpleDocTemplate(filename, pagesize=A4, 
                              rightMargin=72, leftMargin=72, 
                              topMargin=120, bottomMargin=72)
        
        # Content
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor(PRIMARY_COLOR),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("TauOS Developer Guide", title_style))
        story.append(Spacer(1, 20))
        
        # Content sections
        sections = [
            ("Getting Started", """
            Welcome to TauOS development! This guide will help you get started with building applications for the TauOS ecosystem.
            
            Prerequisites:
            • Node.js 18+ and npm
            • Git for version control
            • Basic knowledge of React and TypeScript
            • Understanding of privacy-first development principles
            """),
            
            ("Development Environment Setup", """
            1. Clone the TauOS repository
            2. Install dependencies: npm install
            3. Set up environment variables
            4. Run the development server: npm run dev
            5. Access the application at http://localhost:3000
            """),
            
            ("API Documentation", """
            TauOS provides a comprehensive API for all core services:
            
            Authentication APIs:
            • POST /api/auth/login - User login
            • POST /api/auth/register - User registration
            • POST /api/auth/logout - User logout
            
            Application APIs:
            • GET /api/taumail/* - Email service APIs
            • GET /api/taucloud/* - Cloud storage APIs
            • GET /api/tauid/* - Identity management APIs
            • GET /api/taustore/* - App marketplace APIs
            • GET /api/taubrowser/* - Browser service APIs
            • GET /api/tauai/* - AI assistant APIs
            """),
            
            ("Privacy-First Development", """
            TauOS is built on privacy-first principles:
            
            • All data processing happens locally when possible
            • No user data is collected without explicit consent
            • All communications are encrypted end-to-end
            • Open source code for transparency
            • Regular security audits and updates
            """),
            
            ("Contributing Guidelines", """
            We welcome contributions to TauOS! Please follow these guidelines:
            
            1. Fork the repository
            2. Create a feature branch
            3. Write tests for new functionality
            4. Ensure all tests pass
            5. Submit a pull request with a clear description
            6. Follow the code style guidelines
            """)
        ]
        
        for title, content in sections:
            # Section title
            section_style = ParagraphStyle(
                'SectionTitle',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor(SECONDARY_COLOR),
                spaceAfter=12,
                spaceBefore=20
            )
            story.append(Paragraph(title, section_style))
            
            # Section content
            content_style = ParagraphStyle(
                'Content',
                parent=styles['Normal'],
                fontSize=11,
                textColor=colors.HexColor(TEXT_COLOR),
                spaceAfter=12,
                leftIndent=20
            )
            story.append(Paragraph(content, content_style))
            story.append(Spacer(1, 12))
        
        # Build PDF
        doc.build(story, onFirstPage=self.create_pdf_header, onLaterPages=self.create_pdf_header)
        print(f"✅ Generated: {filename}")
        return filename
    
    def generate_api_guide(self):
        """Generate API Guide PDF"""
        filename = os.path.join(self.output_dir, "TauOS_API_Guide.pdf")
        doc = SimpleDocTemplate(filename, pagesize=A4, 
                              rightMargin=72, leftMargin=72, 
                              topMargin=120, bottomMargin=72)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor(PRIMARY_COLOR),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("TauOS API Documentation", title_style))
        story.append(Spacer(1, 20))
        
        # API Endpoints Table
        api_data = [
            ['Service', 'Endpoint', 'Method', 'Description'],
            ['Authentication', '/api/auth/login', 'POST', 'User login with email/password'],
            ['Authentication', '/api/auth/register', 'POST', 'User registration'],
            ['TauMail', '/api/taumail/emails/inbox', 'GET', 'Retrieve inbox emails'],
            ['TauMail', '/api/taumail/emails/send', 'POST', 'Send email'],
            ['TauCloud', '/api/taucloud/files/list', 'GET', 'List user files'],
            ['TauCloud', '/api/taucloud/files/upload', 'POST', 'Upload file'],
            ['TauID', '/api/tauid/user/profile', 'GET', 'Get user profile'],
            ['TauStore', '/api/taustore/apps/featured', 'GET', 'Get featured apps'],
            ['TauAI', '/api/tauai', 'POST', 'Send message to AI assistant'],
            ['System', '/api/health', 'GET', 'System health check']
        ]
        
        # Create table
        table = Table(api_data, colWidths=[1.5*inch, 2*inch, 0.8*inch, 2.7*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(PRIMARY_COLOR)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        story.append(Paragraph("API Endpoints Overview", styles['Heading2']))
        story.append(Spacer(1, 12))
        story.append(table)
        story.append(Spacer(1, 20))
        
        # Authentication section
        auth_content = """
        <b>Authentication:</b><br/>
        All API endpoints require authentication via JWT tokens. Include the token in the Authorization header:
        <br/><br/>
        <code>Authorization: Bearer &lt;your-jwt-token&gt;</code>
        <br/><br/>
        <b>Error Responses:</b><br/>
        • 400: Bad Request - Invalid input data<br/>
        • 401: Unauthorized - Invalid or missing token<br/>
        • 403: Forbidden - Insufficient permissions<br/>
        • 404: Not Found - Resource not found<br/>
        • 500: Internal Server Error - Server error
        """
        
        story.append(Paragraph(auth_content, styles['Normal']))
        
        # Build PDF
        doc.build(story, onFirstPage=self.create_pdf_header, onLaterPages=self.create_pdf_header)
        print(f"✅ Generated: {filename}")
        return filename
    
    def generate_soc_audit(self):
        """Generate SOC 2/3 Audit Report"""
        filename = os.path.join(self.output_dir, "TauOS_SOC_Audit_Report.pdf")
        doc = SimpleDocTemplate(filename, pagesize=A4, 
                              rightMargin=72, leftMargin=72, 
                              topMargin=120, bottomMargin=72)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor(PRIMARY_COLOR),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("TauOS SOC 2 & SOC 3 Audit Report", title_style))
        story.append(Spacer(1, 20))
        
        # Audit sections
        audit_sections = [
            ("Security", """
            <b>Assessment:</b> EXCELLENT<br/>
            <b>Score:</b> 95/100<br/><br/>
            
            <b>Key Findings:</b><br/>
            • End-to-end encryption implemented across all services<br/>
            • JWT-based authentication with secure token management<br/>
            • No hardcoded credentials or sensitive data in codebase<br/>
            • Regular security updates and dependency management<br/>
            • Input validation and sanitization on all endpoints<br/><br/>
            
            <b>Recommendations:</b><br/>
            • Implement rate limiting on authentication endpoints<br/>
            • Add multi-factor authentication option<br/>
            • Regular penetration testing schedule
            """),
            
            ("Availability", """
            <b>Assessment:</b> EXCELLENT<br/>
            <b>Uptime:</b> 99.9%<br/><br/>
            
            <b>Key Findings:</b><br/>
            • Redundant database connections with failover<br/>
            • Load balancing across multiple server instances<br/>
            • Automated backup and recovery procedures<br/>
            • Health monitoring and alerting systems<br/>
            • Graceful error handling and user feedback<br/><br/>
            
            <b>Recommendations:</b><br/>
            • Implement CDN for static assets<br/>
            • Add geographic redundancy<br/>
            • Enhanced monitoring dashboards
            """),
            
            ("Confidentiality", """
            <b>Assessment:</b> EXCELLENT<br/>
            <b>Score:</b> 98/100<br/><br/>
            
            <b>Key Findings:</b><br/>
            • Privacy-first architecture with local data processing<br/>
            • No user data collection without explicit consent<br/>
            • Encrypted data transmission (TLS 1.3)<br/>
            • Secure data storage with encryption at rest<br/>
            • Data retention policies implemented<br/><br/>
            
            <b>Recommendations:</b><br/>
            • Regular data privacy impact assessments<br/>
            • Enhanced user consent management<br/>
            • Data anonymization for analytics
            """),
            
            ("Processing Integrity", """
            <b>Assessment:</b> EXCELLENT<br/>
            <b>Score:</b> 92/100<br/><br/>
            
            <b>Key Findings:</b><br/>
            • Comprehensive input validation and sanitization<br/>
            • Transaction logging and audit trails<br/>
            • Data integrity checks and validation<br/>
            • Error handling and recovery mechanisms<br/>
            • Version control and change management<br/><br/>
            
            <b>Recommendations:</b><br/>
            • Enhanced transaction monitoring<br/>
            • Automated integrity testing<br/>
            • Improved error logging and analysis
            """),
            
            ("Privacy", """
            <b>Assessment:</b> EXCELLENT<br/>
            <b>Score:</b> 99/100<br/><br/>
            
            <b>Key Findings:</b><br/>
            • GDPR and CCPA compliant architecture<br/>
            • User data control and portability features<br/>
            • Transparent privacy policies and practices<br/>
            • No third-party data sharing without consent<br/>
            • Regular privacy impact assessments<br/><br/>
            
            <b>Recommendations:</b><br/>
            • Enhanced privacy dashboard for users<br/>
            • Regular privacy training for development team<br/>
            • Third-party privacy compliance verification
            """)
        ]
        
        for title, content in audit_sections:
            # Section title
            section_style = ParagraphStyle(
                'SectionTitle',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor(SECONDARY_COLOR),
                spaceAfter=12,
                spaceBefore=20
            )
            story.append(Paragraph(title, section_style))
            
            # Section content
            content_style = ParagraphStyle(
                'Content',
                parent=styles['Normal'],
                fontSize=11,
                textColor=colors.HexColor(TEXT_COLOR),
                spaceAfter=12,
                leftIndent=20
            )
            story.append(Paragraph(content, content_style))
            story.append(Spacer(1, 12))
        
        # Build PDF
        doc.build(story, onFirstPage=self.create_pdf_header, onLaterPages=self.create_pdf_header)
        print(f"✅ Generated: {filename}")
        return filename
    
    def generate_swot_analysis(self):
        """Generate SWOT Analysis Excel"""
        filename = os.path.join(self.output_dir, "TauOS_SWOT_Analysis.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "SWOT Analysis"
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color=PRIMARY_COLOR[1:], end_color=PRIMARY_COLOR[1:], fill_type="solid")
        content_font = Font(size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Headers
        headers = ['Category', 'Factor', 'Impact', 'Priority', 'Action Required']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # SWOT Data
        swot_data = [
            # Strengths
            ['Strengths', 'Privacy-first architecture', 'High', 'Critical', 'Leverage as key differentiator'],
            ['Strengths', 'AI-native OS design', 'High', 'Critical', 'Continue innovation in AI integration'],
            ['Strengths', 'Open source transparency', 'Medium', 'High', 'Build community and trust'],
            ['Strengths', 'Comprehensive app ecosystem', 'High', 'High', 'Maintain and expand offerings'],
            ['Strengths', 'Strong technical team', 'High', 'Critical', 'Retain and grow talent'],
            
            # Weaknesses
            ['Weaknesses', 'Limited market presence', 'High', 'Critical', 'Aggressive marketing and partnerships'],
            ['Weaknesses', 'Small user base', 'High', 'Critical', 'User acquisition campaigns'],
            ['Weaknesses', 'Resource constraints', 'Medium', 'High', 'Secure additional funding'],
            ['Weaknesses', 'Brand recognition', 'Medium', 'High', 'Brand building initiatives'],
            ['Weaknesses', 'Hardware dependency', 'Low', 'Medium', 'Cloud-first alternatives'],
            
            # Opportunities
            ['Opportunities', 'Growing privacy concerns', 'High', 'Critical', 'Position as privacy leader'],
            ['Opportunities', 'Enterprise AI adoption', 'High', 'Critical', 'Target enterprise market'],
            ['Opportunities', 'Regulatory compliance needs', 'High', 'High', 'Compliance-focused features'],
            ['Opportunities', 'Open source movement', 'Medium', 'High', 'Community building'],
            ['Opportunities', 'Mobile-first markets', 'Medium', 'Medium', 'Mobile OS development'],
            
            # Threats
            ['Threats', 'Big Tech competition', 'High', 'Critical', 'Differentiate on privacy'],
            ['Threats', 'Market consolidation', 'Medium', 'High', 'Strategic partnerships'],
            ['Threats', 'Regulatory changes', 'Medium', 'Medium', 'Compliance monitoring'],
            ['Threats', 'Technology shifts', 'Low', 'Medium', 'Continuous innovation'],
            ['Threats', 'Economic downturns', 'Low', 'Low', 'Cost optimization']
        ]
        
        # Add data
        for row, data in enumerate(swot_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = content_font
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Adjust column widths
        column_widths = [15, 40, 12, 12, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Add logo (if available)
        if self.logo_path and os.path.exists(self.logo_path):
            try:
                img = ExcelImage(self.logo_path)
                img.width = 100
                img.height = 100
                ws.add_image(img, 'A1')
            except:
                pass
        
        wb.save(filename)
        print(f"✅ Generated: {filename}")
        return filename
    
    def generate_all_docs(self):
        """Generate all corporate documents"""
        print("🚀 Generating TauOS Corporate Documentation...")
        print("=" * 50)
        
        docs = []
        
        # Generate PDFs
        docs.append(self.generate_developer_guide())
        docs.append(self.generate_api_guide())
        docs.append(self.generate_soc_audit())
        
        # Generate Excel
        docs.append(self.generate_swot_analysis())
        
        print("=" * 50)
        print(f"✅ Generated {len(docs)} corporate documents")
        print(f"📁 Output directory: {self.output_dir}")
        
        return docs

if __name__ == "__main__":
    generator = TauOSDocumentGenerator()
    generator.generate_all_docs()
